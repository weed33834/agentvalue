"""
多模态附件抽取器
每个抽取器负责一种附件类型，将非结构化附件转为文本。
抽取器采用可插拔设计：未配置对应能力时优雅降级，记录占位说明。
"""

import base64
import asyncio
import csv
import io
import logging
from abc import ABC, abstractmethod
from typing import Any, Callable, Dict, List, Optional

logger = logging.getLogger(__name__)


def _get_attachment_payload(att: Dict[str, Any]) -> Optional[bytes]:
    """从附件 dict 中提取二进制内容（优先 data > path > key > url）"""
    import os

    from core.config import get_settings

    # base64 内联数据
    data = att.get("data")
    if data:
        try:
            if isinstance(data, str):
                return base64.b64decode(data)
            if isinstance(data, (bytes, bytearray)):
                return bytes(data)
        except Exception as e:
            logger.warning("附件 base64 解码失败: %s", e)
            return None

    # 本地路径（防止路径遍历：仅允许在 attachment_dir 白名单目录内读取）
    path = att.get("path")
    if path:
        allowed_dir = os.path.realpath(get_settings().attachment_dir)
        real_path = os.path.realpath(path)
        if os.path.commonpath([real_path, allowed_dir]) != allowed_dir:
            logger.warning(
                "附件路径越权访问被拒绝: %s (允许目录: %s)", path, allowed_dir
            )
            return None
        if os.path.exists(real_path):
            try:
                with open(real_path, "rb") as f:
                    return f.read()
            except Exception as e:
                logger.warning("附件文件读取失败 %s: %s", real_path, e)
                return None

    # 对象存储 key：通过 storage 抽象下载（生产路径，前端上传后传 key）
    key = att.get("key")
    if key:
        try:
            from core.storage import get_storage

            return get_storage().download(key)
        except Exception as e:
            logger.warning("附件从对象存储下载失败 key=%s: %s", key, e)
            return None

    # URL（仅记录，不在抽取器内下载，避免阻塞；由上层预处理）
    url = att.get("url")
    if url:
        return None
    return None


# 各文件类型对应的 magic bytes 签名常量，供 _validate_magic_bytes 使用
_MAGIC_PNG = b"\x89PNG\r\n\x1a\n"
_MAGIC_JPEG = b"\xff\xd8\xff"
_MAGIC_RIFF = b"RIFF"
_MAGIC_MP3_ID3 = b"ID3"
_MAGIC_MP3_FB = b"\xff\xfb"
_MAGIC_FTYP = b"ftyp"


def _validate_magic_bytes(data: bytes, expected_types: List[str]) -> bool:
    """校验二进制数据是否符合期望的文件类型签名（基于 magic bytes）。

    expected_types 支持的取值（不区分大小写）:
      - png:  \\x89PNG\\r\\n\\x1a\\n
      - jpeg: \\xff\\xd8\\xff
      - webp: RIFF....WEBP
      - mp3:  ID3 或 \\xff\\xfb
      - wav:  RIFF....WAVE
      - mp4 / m4a (默认): ftyp box（偏移 4~8 处为 'ftyp'）

    任意一个期望类型命中即返回 True；都不匹配或入参为空返回 False。
    """
    if not data or not expected_types:
        return False

    for raw_type in expected_types:
        t = (raw_type or "").lower()
        if t == "png" and data.startswith(_MAGIC_PNG):
            return True
        if t == "jpeg" and data.startswith(_MAGIC_JPEG):
            return True
        if t == "webp":
            if (
                data.startswith(_MAGIC_RIFF)
                and len(data) >= 12
                and data[8:12] == b"WEBP"
            ):
                return True
        if t == "mp3":
            if data.startswith(_MAGIC_MP3_ID3) or data.startswith(_MAGIC_MP3_FB):
                return True
        if t == "wav":
            if (
                data.startswith(_MAGIC_RIFF)
                and len(data) >= 12
                and data[8:12] == b"WAVE"
            ):
                return True
        if t in ("mp4", "m4a"):
            # ISO BMFF ftyp box：偏移 4~8 处为 'ftyp'
            if len(data) >= 8 and data[4:8] == _MAGIC_FTYP:
                return True
    return False


def _detect_kind(mime: str, filename: str) -> str:
    """根据 mime / 文件名推断附件类型：text / image / audio / table / unknown"""
    mime = (mime or "").lower()
    fname = (filename or "").lower()

    # 表格类优先判断（text/csv 也属于表格，避免被 text/* 吞掉）
    if fname.endswith((".csv", ".tsv")) or mime in (
        "text/csv",
        "text/tab-separated-values",
    ):
        return "table"
    if fname.endswith((".xlsx", ".xls")) or mime in (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    ):
        return "table"
    if fname.endswith(".pdf") or mime == "application/pdf":
        return "pdf"
    if mime.startswith("text/") or fname.endswith((".txt", ".md", ".log")):
        return "text"
    if mime.startswith("image/") or fname.endswith(
        (".png", ".jpg", ".jpeg", ".gif", ".webp")
    ):
        return "image"
    if mime.startswith("audio/") or fname.endswith((".wav", ".mp3", ".m4a", ".flac")):
        return "audio"
    if mime.startswith("video/") or fname.endswith((".mp4", ".mov")):
        return "video"
    return "unknown"


class BaseExtractor(ABC):
    """抽取器基类"""

    kind: str = "unknown"

    @abstractmethod
    async def extract(self, attachment: Dict[str, Any]) -> str:
        """抽取文本，返回纯文本"""
        raise NotImplementedError


class TextExtractor(BaseExtractor):
    """文本附件抽取器：直接读取内容"""

    kind = "text"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知文件")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            # 可能直接在 content 字段提供文本
            content = attachment.get("content")
            if content:
                return f"[文本附件 {filename}]\n{content}"
            return f"[文本附件 {filename}] 无法读取内容"
        try:
            text = payload.decode("utf-8", errors="replace")
            return f"[文本附件 {filename}]\n{text}"
        except Exception as e:
            logger.warning("文本附件解码失败 %s: %s", filename, e)
            return f"[文本附件 {filename}] 解码失败"


class TableExtractor(BaseExtractor):
    """表格附件抽取器：解析 CSV/TSV 为 Markdown 表格文本"""

    kind = "table"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知表格")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[表格附件 {filename}] 无法读取内容"

        fname = filename.lower()
        delimiter = "\t" if fname.endswith(".tsv") else ","

        # xlsx 需 openpyxl，未安装时降级
        if fname.endswith((".xlsx", ".xls")):
            try:
                return self._extract_xlsx(payload, filename)
            except ImportError:
                return f"[表格附件 {filename}] xlsx 解析需安装 openpyxl，已跳过"
            except Exception as e:
                logger.warning("xlsx 解析失败 %s: %s", filename, e)
                return f"[表格附件 {filename}] 解析失败: {e}"

        # CSV / TSV
        try:
            text = payload.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text), delimiter=delimiter)
            rows = list(reader)
            if not rows:
                return f"[表格附件 {filename}] 空表格"
            return self._rows_to_markdown(rows, filename)
        except Exception as e:
            logger.warning("CSV 解析失败 %s: %s", filename, e)
            return f"[表格附件 {filename}] 解析失败: {e}"

    def _extract_xlsx(self, payload: bytes, filename: str) -> str:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        parts: List[str] = [f"[表格附件 {filename}]"]
        for sheet in wb.worksheets:
            parts.append(f"## 工作表: {sheet.title}")
            rows = list(sheet.iter_rows(values_only=True))
            if rows:
                parts.append(
                    self._rows_to_markdown(
                        [list(r) for r in rows], filename, header=False
                    )
                )
        wb.close()
        return "\n".join(parts)

    def _rows_to_markdown(
        self, rows: List[List[Any]], filename: str, header: bool = True
    ) -> str:
        if not rows:
            return f"[表格附件 {filename}] 空表格"
        # 截断过宽/过长表格，避免撑爆 prompt
        max_cols = 8
        max_rows = 50
        truncated = rows[:max_rows]
        cols = max(len(r) for r in truncated)
        cols = min(cols, max_cols)

        def _cell(v: Any) -> str:
            if v is None:
                return ""
            return str(v).replace("|", "\\|").replace("\n", " ")

        lines = []
        if header:
            head = [_cell(rows[0][i]) if i < len(rows[0]) else "" for i in range(cols)]
            lines.append("| " + " | ".join(head) + " |")
            lines.append("| " + " | ".join(["---"] * cols) + " |")
            body = truncated[1:]
        else:
            lines.append("| " + " | ".join(["列"] * cols) + " |")
            lines.append("| " + " | ".join(["---"] * cols) + " |")
            body = truncated

        for row in body:
            cells = [_cell(row[i]) if i < len(row) else "" for i in range(cols)]
            lines.append("| " + " | ".join(cells) + " |")

        if len(rows) > max_rows:
            lines.append(f"\n(表格共 {len(rows)} 行，已截断显示前 {max_rows} 行)")
        return "\n".join(lines)


class ImageExtractor(BaseExtractor):
    """
    图片附件抽取器：OCR / 视觉理解。
    优先级：注入的 vision_callable（云端多模态模型）> ocr_extractor（本地/云端 OCR）> 占位降级。
    """

    kind = "image"

    def __init__(
        self,
        vision_callable=None,
        ocr_extractor: Optional["OCRExtractor"] = None,
    ):
        self._vision = vision_callable
        self._ocr = ocr_extractor

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知图片")
        # 1) 云端视觉模型优先（视觉理解，区别于 OCR 的纯文字识别）
        if self._vision is not None:
            payload = _get_attachment_payload(attachment)
            if payload is None:
                return f"[图片附件 {filename}] 无法读取图片数据"
            try:
                b64 = base64.b64encode(payload).decode("ascii")
                # vision_callable 约定签名: (prompt: str, image_data: str) -> str
                # image_data 为 base64 编码的图片二进制
                description = await self._vision(
                    prompt=f"请描述图片内容并提取关键信息(文件名:{filename})",
                    image_data=b64,
                )
                if description:
                    return f"[图片附件 {filename}]\n{description}"
                # vision 返回空(None)表示降级,继续走 OCR
            except Exception as e:
                logger.warning("图片视觉理解失败 %s: %s", filename, e)
                # 不在此 return,继续走 OCR 降级,避免视觉失败时图片内容全丢
        # 2) OCR 抽取器（本地 tesseract / 云端 OCR）
        if self._ocr is not None:
            return await self._ocr.extract(attachment)
        # 3) 未配置任何能力，占位降级
        return f"[图片附件 {filename}] 未配置 OCR/视觉模型，已跳过抽取。请在系统中配置多模态模型以启用图片理解。"


class OCRExtractor(BaseExtractor):
    """OCR 抽取器基类：图片转文本，子类实现具体后端（tesseract / 云端）。"""

    kind = "image"

    @abstractmethod
    async def extract(self, attachment: Dict[str, Any]) -> str:
        raise NotImplementedError


class LocalTesseractOCR(OCRExtractor):
    """
    本地 Tesseract OCR：调 pytesseract。
    tesseract 系统二进制缺失时降级为提示文本而非抛异常，保证系统不崩。
    """

    kind = "image"

    def __init__(self, lang: str = "chi_sim+eng"):
        self._lang = lang

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知图片")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[图片附件 {filename}] 无法读取图片数据"
        try:
            import pytesseract  # type: ignore
        except ImportError:
            return f"[图片附件 {filename}] [OCR 不可用,需安装 pytesseract]"

        image_obj, tmp_path = self._to_image(payload)
        if image_obj is None:
            return f"[图片附件 {filename}] [OCR 不可用,需安装 tesseract]"
        try:
            # P1-阻塞: tesseract 是系统二进制调用,丢线程池避免阻塞事件循环
            text = await asyncio.to_thread(
                pytesseract.image_to_string, image_obj, lang=self._lang
            )
        except Exception as e:
            # tesseract 系统二进制缺失或调用失败：降级不抛
            logger.warning("Tesseract OCR 不可用 %s: %s", filename, e)
            return f"[图片附件 {filename}] [OCR 不可用,需安装 tesseract]"
        finally:
            if tmp_path:
                import os

                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        text = (text or "").strip()
        confidence = await asyncio.to_thread(
            self._mean_confidence, pytesseract, image_obj
        )
        review = self._low_confidence_flag(confidence)
        # P0-泄漏: 释放 PIL Image 文件描述符
        close = getattr(image_obj, "close", None)
        if callable(close):
            try:
                close()
            except Exception:
                pass
        if not text:
            base = f"[图片附件 {filename}]\n(OCR 未识别到文本)"
        else:
            base = f"[图片附件 {filename}]\n{text}"
        if review:
            base += f"\n{review}"
        return base

    def _to_image(self, payload: bytes):
        """字节流转成 tesseract 可消费的图像对象；无 PIL 时落临时文件路径。"""
        try:
            from PIL import Image  # type: ignore

            return Image.open(io.BytesIO(payload)), None
        except Exception:
            pass
        # PIL 不可用 / 图片解码失败：写临时文件，pytesseract 也能接受文件路径
        import os
        import tempfile

        try:
            fd, path = tempfile.mkstemp(suffix=".img")
            with os.fdopen(fd, "wb") as f:
                f.write(payload)
            return path, path
        except Exception:
            return None, None

    def _mean_confidence(self, pytesseract, image_obj) -> Optional[float]:
        """best-effort 取 OCR 平均置信度（0~1），失败返回 None。"""
        try:
            from pytesseract import Output  # type: ignore

            data = pytesseract.image_to_data(
                image_obj, lang=self._lang, output_type=Output.DICT
            )
            confs = [
                int(c)
                for c in data.get("conf", [])
                if str(c).lstrip("-").isdigit() and int(c) >= 0
            ]
            if not confs:
                return None
            return sum(confs) / len(confs) / 100.0
        except Exception:
            return None

    def _low_confidence_flag(self, confidence: Optional[float]) -> str:
        """置信度低于阈值时返回人工复核标记，否则空串。"""
        if confidence is None:
            return ""
        try:
            from core.config import get_settings

            threshold = get_settings().multimodal_confidence_threshold
        except Exception:
            threshold = 0.7
        if confidence < threshold:
            return f"[置信度 {confidence:.2f} 低于阈值 {threshold},建议人工复核]"
        return ""


class CloudOCR(OCRExtractor):
    """
    云端 OCR：基于 OpenAI 兼容 vision API（gpt-4o / gpt-4o-mini / qwen-vl-plus / glm-4v 等）。

    两条调用路径：
    1. 优先使用注入的 vision_callable（从 ModelRouter 拿 provider，避免重复构造 client）；
    2. 退化到内置 AsyncOpenAI client，按 api_key/base_url/model 直连 vision 接口。

    未配置 api_key 且未注入 vision_callable 时返回占位提示并标记建议人工复核，
    由上层（MultimodalCleaner / 调用方）决定是否降级到 tesseract。
    """

    kind = "image"

    def __init__(
        self,
        provider: str = "openai",
        api_key: Optional[str] = None,
        secret_key: Optional[str] = None,
        base_url: Optional[str] = None,
        model: str = "gpt-4o-mini",
        vision_callable: Optional[Callable] = None,
    ):
        self._provider = provider
        self._api_key = api_key
        self._secret_key = secret_key
        self._base_url = base_url
        self._model = model
        self._vision_callable = vision_callable

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知图片")
        if not self._api_key and not self._vision_callable:
            return (
                f"[图片附件 {filename}] [云端 OCR 未配置 API Key,已跳过;建议人工复核]"
            )
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[图片附件 {filename}] 无法读取图片数据"
        # H9: magic bytes 校验，防止 base64 编码后非图片二进制被送进 vision 接口
        if not _validate_magic_bytes(payload, ["png", "jpeg", "webp"]):
            raise ValueError(
                f"图片附件 {filename} 不是合法的 PNG/JPEG/WebP 格式 (mime={attachment.get('mime', 'unknown')})"
            )
        try:
            if self._vision_callable:
                # 注入路径：从 ModelRouter 拿 provider，签名 (prompt, image_data) -> str
                text = await self._vision_callable(
                    prompt="请识别图片中所有文字,按原排版输出",
                    image_data=base64.b64encode(payload).decode("ascii"),
                )
            else:
                # 直连路径：OpenAI 兼容 vision chat completions
                from openai import AsyncOpenAI

                b64 = base64.b64encode(payload).decode("ascii")
                # P0-泄漏: 用 async with 确保 HTTP 连接池释放
                async with AsyncOpenAI(
                    api_key=self._api_key, base_url=self._base_url
                ) as client:
                    response = await client.chat.completions.create(
                        model=self._model,
                        messages=[
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "text",
                                        "text": "请识别图片中所有文字,按原排版输出,不要加注释",
                                    },
                                    {
                                        "type": "image_url",
                                        "image_url": {
                                            "url": f"data:image/jpeg;base64,{b64}"
                                        },
                                    },
                                ],
                            }
                        ],
                        max_tokens=1500,
                    )
                    text = response.choices[0].message.content or ""
        except Exception as e:
            logger.warning("云端 OCR 调用失败 %s: %s", filename, e)
            return f"[图片附件 {filename}] [云端 OCR 调用失败: {e};建议人工复核]"

        text = (text or "").strip()
        if not text:
            return f"[图片附件 {filename}]\n(云端 OCR 未识别到文本;建议人工复核)"
        # vision 模型无法直接给置信度，假定 0.85；输出过短可能漏识别，标记复核
        review_tag = " [输出过短,建议人工复核]" if len(text) < 20 else ""
        return f"[图片附件 {filename}]\n{text}{review_tag}"


class AudioExtractor(BaseExtractor):
    """
    音频附件抽取器：ASR 语音转文字。
    优先使用注入的 asr_callable（如 Whisper）；未配置时降级为占位说明。
    """

    kind = "audio"

    def __init__(self, asr_callable=None):
        self._asr = asr_callable

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知音频")
        if self._asr is None:
            return f"[音频附件 {filename}] 未配置 ASR 语音识别模型，已跳过抽取。请在系统中配置 ASR 模型以启用语音转文字。"
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[音频附件 {filename}] 无法读取音频数据"
        try:
            transcript = await self._asr(payload, filename)
            return f"[音频附件 {filename}]\n{transcript}"
        except Exception as e:
            logger.warning("音频抽取失败 %s: %s", filename, e)
            return f"[音频附件 {filename}] 抽取失败: {e}"


class WhisperASR(AudioExtractor):
    """
    Whisper ASR：基于 OpenAI 兼容 audio transcription API（whisper-1）。

    Phase 10 重构：不再依赖本地 openai-whisper 重模型，改走云端 whisper-1 接口。
    未配置 asr_api_key 时返回占位提示并标记建议人工复核，
    由上层（MultimodalCleaner / 调用方）决定是否降级到 DummyASR。
    """

    kind = "audio"

    def __init__(
        self,
        model_name: str = "base",
        api_key: Optional[str] = None,
        base_url: Optional[str] = None,
        model: str = "whisper-1",
    ):
        super().__init__(asr_callable=None)
        # model_name 保留向后兼容（本地 whisper 模型档位），实际不再使用
        self._model_name = model_name
        self._api_key = api_key
        self._base_url = base_url
        self._model = model

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知音频")
        if not self._api_key:
            return f"[音频附件 {filename}] [Whisper ASR 未配置 api_key,已跳过;建议人工复核]"
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[音频附件 {filename}] 无法读取音频数据"
        # H9: magic bytes 校验，防止非音频二进制被送进 whisper transcription 接口
        if not _validate_magic_bytes(payload, ["mp3", "wav", "mp4", "m4a"]):
            raise ValueError(
                f"音频附件 {filename} 不是合法的 MP3/WAV/MP4/M4A 格式 (mime={attachment.get('mime', 'unknown')})"
            )
        try:
            from openai import AsyncOpenAI

            # attachment.content 是 bytes，需包装成文件对象供 SDK 上传
            audio_file = io.BytesIO(payload)
            audio_file.name = filename or "audio.mp3"
            # P0-泄漏: 用 async with 确保 HTTP 连接池释放
            async with AsyncOpenAI(
                api_key=self._api_key, base_url=self._base_url
            ) as client:
                response = await client.audio.transcriptions.create(
                    model=self._model,
                    file=audio_file,
                )
                text = response.text or ""
        except Exception as e:
            logger.warning("Whisper ASR 调用失败 %s: %s", filename, e)
            return f"[音频附件 {filename}] [Whisper ASR 调用失败: {e};建议人工复核]"

        text = (text or "").strip()
        if not text:
            return f"[音频附件 {filename}]\n(Whisper ASR 未识别到文本;建议人工复核)"
        # Whisper 通常较高置信度（0.9）；输出过短可能漏识别，标记复核
        review_tag = " [输出过短,建议人工复核]" if len(text) < 10 else ""
        return f"[音频附件 {filename}]\n{text}{review_tag}"


class PdfExtractor(BaseExtractor):
    """
    PDF 附件抽取器：提取文本层。
    生产默认走 pypdf（纯 Python、无系统依赖，MVP 选型）；pdfplumber 若存在则优先用，
    对扫描件/复杂版式兼容更好。两者都不可用时降级。
    """

    kind = "pdf"
    # 最多抽取前 N 页，避免超大 PDF 撑爆 prompt
    MAX_PAGES = 20

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知PDF")
        payload = _get_attachment_payload(attachment)
        if payload is None:
            return f"[PDF附件 {filename}] 无法读取数据"
        # P1-阻塞: PDF 解析是 CPU/IO 密集同步操作,丢线程池避免阻塞事件循环
        return await asyncio.to_thread(self._extract_pdf, payload, filename)

    def _extract_pdf(self, payload: bytes, filename: str) -> str:
        # pdfplumber 优先：兼容更广，部分残缺/复杂版式 PDF 表现更好
        try:
            import pdfplumber  # type: ignore
        except ImportError:
            pdfplumber = None  # type: ignore

        if pdfplumber is not None:
            try:
                pages_text: List[str] = []
                with pdfplumber.open(io.BytesIO(payload)) as pdf:
                    for i, page in enumerate(pdf.pages[: self.MAX_PAGES]):
                        text = page.extract_text() or ""
                        if text:
                            pages_text.append(f"--- 第 {i + 1} 页 ---\n{text}")
                body = (
                    "\n\n".join(pages_text)
                    if pages_text
                    else "(无可提取文本，可能是扫描件)"
                )
                return f"[PDF附件 {filename}]\n{body}"
            except Exception as e:
                # pdfplumber 自身解析失败时回退 pypdf，不直接放弃
                logger.warning("pdfplumber 解析失败 %s: %s，回退 pypdf", filename, e)

        # pypdf 回退：纯 Python 无系统依赖，MVP 默认依赖
        try:
            from pypdf import PdfReader  # type: ignore
        except ImportError:
            return f"[PDF附件 {filename}] PDF 解析需安装 pdfplumber 或 pypdf，已跳过"

        try:
            reader = PdfReader(io.BytesIO(payload))
        except Exception as e:
            # 解析失败兜底：引导安装兼容性更好的 pdfplumber（对残缺 PDF 更鲁棒）
            logger.warning("PDF 读取失败 %s: %s", filename, e)
            return f"[PDF附件 {filename}] 解析失败，可尝试安装 pdfplumber 获得更好兼容性: {e}"

        # 加密 PDF 检测：给出清晰错误，而非吞掉
        if getattr(reader, "is_encrypted", False):
            return f"[PDF附件 {filename}] PDF 已加密，需提供密码或解除加密后重试"

        pages_text = []
        for i, page in enumerate(reader.pages[: self.MAX_PAGES]):
            try:
                text = page.extract_text() or ""
            except Exception:
                # 单页解析失败不影响其它页
                text = ""
            if text:
                pages_text.append(f"--- 第 {i + 1} 页 ---\n{text}")
        body = "\n\n".join(pages_text) if pages_text else "(无可提取文本，可能是扫描件)"
        return f"[PDF附件 {filename}]\n{body}"


class UnknownExtractor(BaseExtractor):
    """未知类型附件：记录占位"""

    kind = "unknown"

    async def extract(self, attachment: Dict[str, Any]) -> str:
        filename = attachment.get("filename", "未知附件")
        mime = attachment.get("mime", "unknown")
        return f"[附件 {filename}（{mime}）] 暂不支持的附件类型，已跳过"
